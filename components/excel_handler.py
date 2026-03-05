import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import io

DAYS = ["Sunday", "Monday", "Tuesday", "Wednesday", "Thursday", "Friday", "Saturday"]
MEALS = ["Breakfast", "Lunch", "Evening Snack", "Dinner"]


def load_recipes(filepath: str) -> pd.DataFrame:
    """Load recipes from Excel file."""
    df = pd.read_excel(filepath)
    df.columns = df.columns.str.strip()
    return df


def format_recipes_for_prompt(df: pd.DataFrame) -> str:
    """Convert recipes dataframe to a string for the LLM prompt."""
    lines = []
    for _, row in df.iterrows():
        veg = str(row.get("Vegetarian", "")).strip().lower()
        if veg == "yes":
            name = row.get("Recipe Name", "")
            link = row.get("Link/URL", "N/A")
            ingredients = row.get("Ingredients", "")
            category = row.get("Category", "")
            lines.append(f"- {name} ({category}) | Ingredients: {ingredients} | Link: {link}")
    return "\n".join(lines)


def generate_meal_plan_excel(meal_plan: dict) -> bytes:
    """Generate a formatted Excel meal plan and return as bytes."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Weekly Meal Plan"

    # Styles
    header_fill = PatternFill("solid", start_color="2E4057")
    header_font = Font(bold=True, color="FFFFFF", name="Arial", size=11)
    day_fill = PatternFill("solid", start_color="4F81BD")
    day_font = Font(bold=True, color="FFFFFF", name="Arial", size=10)
    meal_label_fill = PatternFill("solid", start_color="DBE5F1")
    meal_label_font = Font(bold=True, name="Arial", size=10)
    cell_font = Font(name="Arial", size=9)
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left = Alignment(horizontal="left", vertical="top", wrap_text=True)
    thin = Side(style="thin", color="BFBFBF")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    # Title
    ws.merge_cells("A1:I1")
    title_cell = ws["A1"]
    title_cell.value = "🍽️ MySousChef — Weekly Meal Plan"
    title_cell.font = Font(bold=True, name="Arial", size=14, color="2E4057")
    title_cell.alignment = center
    title_cell.fill = PatternFill("solid", start_color="E8F0FE")
    ws.row_dimensions[1].height = 30

    # Header row: Meal | Days
    ws.cell(row=2, column=1, value="Meal").font = header_font
    ws.cell(row=2, column=1).fill = header_fill
    ws.cell(row=2, column=1).alignment = center
    ws.cell(row=2, column=1).border = border

    for col, day in enumerate(DAYS, 2):
        cell = ws.cell(row=2, column=col, value=day)
        cell.font = day_font
        cell.fill = day_fill
        cell.alignment = center
        cell.border = border
    ws.row_dimensions[2].height = 20

    # Meal rows
    meal_colors = {
        "Breakfast": "FFF2CC",
        "Lunch": "E2EFDA",
        "Evening Snack": "FCE4D6",
        "Dinner": "DDEBF7"
    }

    for row_idx, meal in enumerate(MEALS, 3):
        # Meal label
        label_cell = ws.cell(row=row_idx, column=1, value=meal)
        label_cell.font = meal_label_font
        label_cell.fill = meal_label_fill
        label_cell.alignment = center
        label_cell.border = border

        for col_idx, day in enumerate(DAYS, 2):
            meal_data = meal_plan.get(day, {}).get(meal, {})
            recipe = meal_data.get("recipe", "—")
            link = meal_data.get("link", "")
            protein = meal_data.get("protein", "")
            fiber = meal_data.get("fiber", "")
            carbs = meal_data.get("carbs", "")

            nutrient_text = f"P:{protein} | F:{fiber} | C:{carbs}" if protein else ""
            cell_value = f"{recipe}\n{nutrient_text}" if nutrient_text else recipe

            cell = ws.cell(row=row_idx, column=col_idx, value=cell_value)
            cell.font = cell_font
            cell.fill = PatternFill("solid", start_color=meal_colors[meal])
            cell.alignment = left
            cell.border = border

            # Add hyperlink if available
            if link and link != "N/A" and link.startswith("http"):
                cell.hyperlink = link
                cell.font = Font(name="Arial", size=9, color="0563C1", underline="single")

        ws.row_dimensions[row_idx].height = 55

    # Column widths
    ws.column_dimensions["A"].width = 16
    for col in range(2, 9):
        ws.column_dimensions[get_column_letter(col)].width = 22

    # Freeze header
    ws.freeze_panes = "B3"

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output.getvalue()
